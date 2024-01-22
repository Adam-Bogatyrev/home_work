import openpyxl

new_workbook = openpyxl.Workbook()
new_worksheet = new_workbook.active
new_worksheet.title = 'page 1'

rows = [[f"{column_index} {row_index}" for column_index in "ABCDEFGH"] for row_index in range(1, 10)]

for row_index, row in enumerate(rows, 1):
    for column_index, value in enumerate(row, 1):
        new_worksheet.cell(row_index, column_index, value)

new_workbook.save("temp/new_data.xlsx")

for row_index in range(1, new_worksheet.max_row + 1):
    for column_index in range(1, new_worksheet.max_column + 1):
        cell = new_worksheet.cell(row=row_index, column=column_index)

print("Файл 'temp/new_data.xlsx' успешно создан.")
